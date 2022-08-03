#В данном скрипте берем из справочника поля Наименование_старое, Наименование_новое, Тип оборудования Новый, Продакт_намбер_новый   
#И меняем их в целевом файле

import openpyxl as pxl
wdict = pxl.load_workbook('C:/Users/m.shcherbina/PycharmProjects/ITOIUItransfer/dictionary_global.xlsx')
wSdict = wdict['переносим']
wSnodict = wdict['не переносим']
wbook = pxl.load_workbook('C:/Users/m.shcherbina/PycharmProjects/ITOIUItransfer/For_powershell_task_3_output.xlsx')
wSbook = wbook.active
#wSnobook = wbook.get
print(wdict.sheetnames)

#Создаем словарь из справочника
replacedict = {}
noreplacedict = {}
for i in range(1, wSdict.max_row + 1):
    name_old_id  = wSdict.cell(i, 3).value
    name_new = wSdict.cell(i, 2).value
    type_ob_new = wSdict.cell(i, 4).value
    product_num_new =  wSdict.cell(i, 6).value
    replacedict[name_old_id] = {'name': name_new, 'type_ob': type_ob_new, 'product': product_num_new }
for i in range(1, wSnodict.max_row + 1):
    name_old_id  = wSnodict.cell(i, 3).value
    name_new = wSnodict.cell(i, 2).value
    type_ob_new = wSnodict.cell(i, 4).value
    product_num_new =  wSnodict.cell(i, 6).value
    noreplacedict[name_old_id] = {'name': name_new, 'type_ob': type_ob_new, 'product': product_num_new }



#print(replacedict)
print('Загружено из справочника', wSdict.max_row, ' значений')
j=0
n=0
k=0
for i in range(1, wSbook.max_row + 1):
   for key in replacedict:
        if wSbook.cell(i, 8).value == key:
            new_name = replacedict[key]['name']
            new_type = replacedict[key]['type_ob']
            new_product = replacedict[key]['product']
            wSbook.cell(i, 8).value = str(new_name) + ', '+ str(wSbook.cell(i, 9).value) #формируем новое наименование из типа имущества и серийника
            wSbook.cell(i, 2).value = new_type
            wSbook.cell(i, 9).value = new_product
            wSbook.cell(i, 1).value = 'Переносим в ИУИ'
            j += 1
            print('Выполнена замена для', key,  'в строке', i, 'Имущество помечено как переносимое')

for i in range(1, wSbook.max_row + 1):
   for key in noreplacedict:
        if wSbook.cell(i, 8).value == key:
            new_name = noreplacedict[key]['name']
            new_type = noreplacedict[key]['type_ob']
            new_product = noreplacedict[key]['product']
            #wSbook.cell(i, 8).value = str(new_name) + ', '+ str(wSbook.cell(i, 9).value) #формируем новое наименование из типа имущества и серийника
            wSbook.cell(i, 2).value = new_type
            #wSbook.cell(i, 9).value = new_product
            wSbook.cell(i, 1).value = 'Не Переносим в ИУИ'
            k += 1
            print( key,  'в строке', i, ' помечено как не переносимое')




wbook.save('C:/Users/m.shcherbina/PycharmProjects/ITOIUItransfer/output.xlsx')
print('Файл сохранен, выполнено', j, 'замен имущества, которое будет перенесено')
print('выполнено', k, 'замен имущества, которое не будет перенесено')
#print('для', n, 'записей справочника соотвествий не найдено')
