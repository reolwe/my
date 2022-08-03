import openpyxl as pxl
wdict = pxl.load_workbook('C:/Users/m.shcherbina/PycharmProjects/ITOIUItransfer/dictionary_simple.xlsx')
wSdict = wdict.active
wbook = pxl.load_workbook('C:/Users/m.shcherbina/PycharmProjects/ITOIUItransfer/For_powershell_task_3_output.xlsx')
wSbook = wbook.active

replacedict = {}
for i in range(1, wSdict.max_row + 1):

    cell_value_class = wSdict.cell(i, 1).value
    cell_value_id = wSdict.cell(i, 2).value
    replacedict[cell_value_id] = cell_value_class
#    define[cell_value_id] = cell_value_class
#    replacedict.append(define)
print(replacedict)
print(wSdict.max_row)
j=0
for i in range(1, wSbook.max_row + 1):
   for key in replacedict:
      if wSbook.cell(i, 1).value == key:
        newCell = replacedict[key]
        wSbook.cell(i, 1).value = newCell
        j += 1
        print('Выполнена замена', key, 'на', replacedict[key], 'в строке', i)
wbook.save('C:/Users/m.shcherbina/PycharmProjects/ITOIUItransfer/output.xlsx')
print('Файл сохранен, выполнено', j, 'замен')